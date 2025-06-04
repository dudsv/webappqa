import os
import re
import time
import json
from flask import Flask, render_template, request, jsonify, send_from_directory, current_app
from werkzeug.utils import secure_filename
import html5lib
from bs4 import BeautifulSoup
import requests
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from docx import Document

# Configuração do aplicativo Flask
app = Flask(__name__, static_folder='static')
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
app.config['RESULTS_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'results')

# Garantir que os diretórios existam
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['RESULTS_FOLDER'], exist_ok=True)

# Funções de processamento de texto
def clean_text(text):
    """Limpa o texto removendo espaços extras e caracteres especiais."""
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    return text

def extract_docx_text(docx_path):
    """Extrai texto de um arquivo DOCX."""
    doc = Document(docx_path)
    texts = []
    for para in doc.paragraphs:
        if para.text.strip():
            texts.append(clean_text(para.text))
    return texts

def load_url_text(url):
    """Carrega texto de uma URL usando Playwright para renderizar JavaScript."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url, wait_until="networkidle")
        
        # Extrair título
        title = page.title()
        
        # Extrair conteúdo HTML
        html_content = page.content()
        
        # Extrair metadados
        metadata = {}
        meta_tags = page.query_selector_all('meta')
        for tag in meta_tags:
            name = tag.get_attribute('name') or tag.get_attribute('property')
            content = tag.get_attribute('content')
            if name and content:
                metadata[name] = content
        
        # Extrair alt tags de imagens
        alt_tags = {}
        images = page.query_selector_all('img')
        for i, img in enumerate(images):
            alt = img.get_attribute('alt')
            src = img.get_attribute('src')
            if alt and src:
                alt_tags[f"Image {i+1}"] = {"alt": alt, "src": src}
        
        # Extrair elementos da página
        elements = []
        
        # Extrair tabela de avaliação veterinária
        vet_tables = page.query_selector_all('.breed-table')
        for table in vet_tables:
            rows = table.query_selector_all('tr')
            for row in rows:
                cells = row.query_selector_all('td')
                if len(cells) >= 2:
                    label = cells[0].inner_text().strip()
                    value = cells[1].inner_text().strip()
                    elements.append(["Puntuación Veterinaria", "table", f"{label}: {value}", ""])
        
        # Extrair outros elementos
        headings = page.query_selector_all('h1, h2, h3, h4, h5, h6')
        for h in headings:
            elements.append([f"Heading {h.evaluate('node => node.tagName').lower()}", h.evaluate('node => node.tagName').lower(), h.inner_text().strip(), ""])
        
        paragraphs = page.query_selector_all('p')
        for p in paragraphs:
            elements.append(["Paragraph", "p", p.inner_text().strip(), ""])
        
        links = page.query_selector_all('a')
        for link in links:
            href = link.get_attribute('href')
            elements.append(["Link", "a", link.inner_text().strip(), href or ""])
        
        # Extrair texto principal
        main_content = page.query_selector('main') or page.query_selector('article') or page.query_selector('body')
        main_text = main_content.inner_text() if main_content else ""
        
        # Extrair todos os textos visíveis
        all_text = page.inner_text('body')
        
        # Fechar o navegador
        browser.close()
        
        # Processar o HTML com BeautifulSoup para extração adicional
        soup = BeautifulSoup(html_content, 'html5lib')
        
        # Extrair textos de elementos específicos
        texts = []
        for element in soup.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'li']):
            text = clean_text(element.get_text())
            if text:
                texts.append(text)
        
        return texts, main_text, metadata, alt_tags, title, elements

def compare_texts(doc_texts, web_texts):
    """Compara textos do documento com textos da web."""
    results = []
    
    for doc_text in doc_texts:
        best_match = None
        best_similarity = 0
        
        for web_text in web_texts:
            # Calcular similaridade simples baseada em palavras comuns
            doc_words = set(doc_text.lower().split())
            web_words = set(web_text.lower().split())
            
            if not doc_words or not web_words:
                continue
                
            common_words = doc_words.intersection(web_words)
            similarity = len(common_words) / max(len(doc_words), len(web_words))
            
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = web_text
        
        # Determinar status baseado na similaridade
        if best_similarity >= 0.9:
            status = "Exato"
        elif best_similarity >= 0.7:
            status = "Similar"
        elif best_similarity >= 0.5:
            status = "Parcial"
        else:
            status = "Não encontrado"
            best_match = None
            best_similarity = None
        
        results.append({
            "doc_text": doc_text,
            "web_text": best_match,
            "status": status,
            "similarity": best_similarity
        })
    
    return results

def generate_summary(comparison_results):
    """Gera um resumo dos resultados da comparação."""
    counts = {
        "exact": 0,
        "similar": 0,
        "partial": 0,
        "missing": 0
    }
    
    for result in comparison_results:
        if result["status"] == "Exato":
            counts["exact"] += 1
        elif result["status"] == "Similar":
            counts["similar"] += 1
        elif result["status"] == "Parcial":
            counts["partial"] += 1
        else:
            counts["missing"] += 1
    
    return counts

def generate_summary_table(comparison_results):
    """Gera uma tabela de resumo para os resultados da comparação."""
    counts = {
        "Exato": 0,
        "Similar": 0,
        "Parcial": 0,
        "Não encontrado": 0
    }
    
    for result in comparison_results:
        counts[result["status"]] += 1
    
    total = sum(counts.values())
    
    summary_table = {}
    for status, count in counts.items():
        percent = (count / total * 100) if total > 0 else 0
        summary_table[status] = {"count": count, "percent": percent}
    
    summary_table["TOTAL"] = {"count": total, "percent": 100}
    
    return summary_table

def create_excel_report(excel_path, comparison_results, summary_table, elements, docx_path, web_url):
    """Cria um relatório Excel com os resultados da comparação."""
    wb = openpyxl.Workbook()
    
    # Criar aba de comparação
    ws_comparison = wb.active
    ws_comparison.title = "Comparacao"
    
    # Adicionar cabeçalhos
    headers = ["Document Text", "Webpage Match", "Status", "Similarity"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_comparison.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Adicionar dados de comparação
    for row_num, result in enumerate(comparison_results, 2):
        ws_comparison.cell(row=row_num, column=1).value = result["doc_text"]
        ws_comparison.cell(row=row_num, column=2).value = result["web_text"] if result["web_text"] else "Não encontrado"
        ws_comparison.cell(row=row_num, column=3).value = result["status"]
        
        if result["similarity"] is not None:
            ws_comparison.cell(row=row_num, column=4).value = f"{result['similarity']:.2%}"
        else:
            ws_comparison.cell(row=row_num, column=4).value = "N/A"
        
        # Adicionar cores baseadas no status
        if result["status"] == "Exato":
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif result["status"] == "Similar":
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif result["status"] == "Parcial":
            fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for col in range(1, 5):
            ws_comparison.cell(row=row_num, column=col).fill = fill
    
    # Ajustar largura das colunas
    for col in range(1, 5):
        ws_comparison.column_dimensions[get_column_letter(col)].width = 40
    
    # Criar aba de resumo
    ws_summary = wb.create_sheet(title="Resumo")
    
    # Adicionar cabeçalhos do resumo
    summary_headers = ["Status", "Quantidade", "Porcentagem"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Adicionar dados do resumo
    row_num = 2
    for status, info in summary_table.items():
        ws_summary.cell(row=row_num, column=1).value = status
        ws_summary.cell(row=row_num, column=2).value = info["count"]
        ws_summary.cell(row=row_num, column=3).value = f"{info['percent']:.2f}%"
        
        # Adicionar cores baseadas no status
        if status == "Exato":
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "Similar":
            fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status == "Parcial":
            fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        elif status == "Não encontrado":
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for col in range(1, 4):
            ws_summary.cell(row=row_num, column=col).fill = fill
        
        row_num += 1
    
    # Ajustar largura das colunas do resumo
    for col in range(1, 4):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20
    
    # Criar aba de elementos da página
    ws_elements = wb.create_sheet(title="Elementos da Página")
    
    # Adicionar cabeçalhos dos elementos
    element_headers = ["Definition", "Tag", "Text", "Link"]
    for col_num, header in enumerate(element_headers, 1):
        cell = ws_elements.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Adicionar dados dos elementos
    for row_num, element in enumerate(elements, 2):
        for col_num, value in enumerate(element, 1):
            ws_elements.cell(row=row_num, column=col_num).value = value
    
    # Ajustar largura das colunas dos elementos
    ws_elements.column_dimensions[get_column_letter(1)].width = 20
    ws_elements.column_dimensions[get_column_letter(2)].width = 10
    ws_elements.column_dimensions[get_column_letter(3)].width = 60
    ws_elements.column_dimensions[get_column_letter(4)].width = 40
    
    # Salvar o arquivo Excel
    wb.save(excel_path)
    return excel_path

# Rotas do aplicativo
@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """
    Processa um único par de documento DOCX e URL para comparação.
    """
    try:
        # Verificar se o arquivo foi enviado
        if 'docx_file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo DOCX enviado'}), 400
        
        docx_file = request.files['docx_file']
        web_url = request.form.get('web_url')
        
        if not docx_file or not web_url:
            return jsonify({'error': 'Arquivo DOCX e URL são obrigatórios'}), 400
        
        # Salvar arquivo temporário
        temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(int(time.time())))
        os.makedirs(temp_dir, exist_ok=True)
        
        docx_path = os.path.join(temp_dir, secure_filename(docx_file.filename))
        docx_file.save(docx_path)
        
        # Processar comparação
        docx_texts = extract_docx_text(docx_path)
        web_texts, main, metadata, alt_tags, title, elements = load_url_text(web_url)
        
        # Calcular similaridades
        comparison_results = compare_texts(docx_texts, web_texts)
        
        # Gerar resumo
        summary = generate_summary(comparison_results)
        summary_table = generate_summary_table(comparison_results)
        
        # Extrair avaliações veterinárias
        vet_ratings = []
        for element in elements:
            if element[0] == 'Puntuación Veterinaria':
                vet_ratings.append(element[2])
        
        # Gerar Excel
        excel_filename = f"comparison_{int(time.time())}.xlsx"
        excel_path = os.path.join(app.config['RESULTS_FOLDER'], excel_filename)
        
        create_excel_report(
            excel_path,
            comparison_results,
            summary_table,
            elements,
            docx_path,
            web_url
        )
        
        # URL para download do Excel
        excel_url = f"/download/{excel_filename}"
        
        return jsonify({
            'success': True,
            'summary': summary,
            'summary_table': summary_table,
            'comparison': [
                {
                    'doc_text': item['doc_text'],
                    'web_text': item['web_text'],
                    'status': item['status'],
                    'similarity': item['similarity']
                }
                for item in comparison_results
            ],
            'elements': [
                {
                    'definition': elem[0],
                    'tag': elem[1],
                    'text': elem[2],
                    'link': elem[3]
                }
                for elem in elements
            ],
            'vet_ratings': vet_ratings,
            'excel_url': excel_url,
            'metadata': metadata
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/batch_upload', methods=['POST'])
def batch_upload():
    """
    Processa múltiplos pares de documento DOCX e URL para comparação em lote.
    Usa a mesma lógica da comparação singular para cada par.
    """
    try:
        pair_count = int(request.form.get('pair_count', 0))
        if pair_count <= 0:
            return jsonify({'error': 'Nenhum par de comparação fornecido'}), 400
        
        results = []
        
        for i in range(pair_count):
            docx_file = request.files.get(f'docx_file_{i}')
            web_url = request.form.get(f'web_url_{i}')
            
            if not docx_file or not web_url:
                continue
            
            # Salvar arquivo temporário
            temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(int(time.time())))
            os.makedirs(temp_dir, exist_ok=True)
            
            docx_path = os.path.join(temp_dir, secure_filename(docx_file.filename))
            docx_file.save(docx_path)
            
            # Processar comparação - mesma lógica da comparação singular
            docx_texts = extract_docx_text(docx_path)
            web_texts, main, metadata, alt_tags, title, elements = load_url_text(web_url)
            
            # Calcular similaridades
            comparison_results = compare_texts(docx_texts, web_texts)
            
            # Gerar resumo
            summary = generate_summary(comparison_results)
            summary_table = generate_summary_table(comparison_results)
            
            # Extrair avaliações veterinárias
            vet_ratings = []
            for element in elements:
                if element[0] == 'Puntuación Veterinaria':
                    vet_ratings.append(element[2])
            
            # Gerar Excel com formatação condicional e cores
            excel_filename = f"comparison_{i}_{int(time.time())}.xlsx"
            excel_path = os.path.join(app.config['RESULTS_FOLDER'], excel_filename)
            
            create_excel_report(
                excel_path,
                comparison_results,
                summary_table,
                elements,
                docx_path,
                web_url
            )
            
            # URL para download do Excel
            excel_url = f"/download/{excel_filename}"
            
            # Adicionar resultado ao lote com informações detalhadas
            results.append({
                'title': f"{os.path.basename(docx_path)} e {web_url}",
                'summary': summary,
                'summary_table': summary_table,
                'comparison': [
                    {
                        'doc_text': item['doc_text'],
                        'web_text': item['web_text'],
                        'status': item['status'],
                        'similarity': item['similarity']
                    }
                    for item in comparison_results
                ],
                'elements': [
                    {
                        'definition': elem[0],
                        'tag': elem[1],
                        'text': elem[2],
                        'link': elem[3]
                    }
                    for elem in elements
                ],
                'vet_ratings': vet_ratings,
                'excel_url': excel_url,
                'metadata': metadata
            })
        
        return jsonify({
            'success': True,
            'results': results
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """
    Rota para download do arquivo Excel com os resultados.
    """
    return send_from_directory(app.config['RESULTS_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=False)
