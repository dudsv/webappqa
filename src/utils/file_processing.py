import os
import pandas as pd
from docx import Document
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def load_docx_text(path):
    """
    Carrega e processa o texto de um arquivo DOCX.
    """
    doc = Document(path)
    return [p.text for p in doc.paragraphs if p.text.strip()]

def save_to_excel(df_compare, df_summary, df_elements, filename):
    """
    Salva os resultados da comparação em um arquivo Excel com formatação de cores.
    
    Args:
        df_compare: DataFrame com os resultados da comparação
        df_summary: DataFrame com o resumo estatístico
        df_elements: DataFrame com os elementos HTML coletados
        filename: Nome do arquivo Excel a ser salvo
    """
    # Cria um escritor Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Salva cada DataFrame em uma aba separada com nomes em português
    df_compare.to_excel(writer, sheet_name='Comparacao', index=False)
    
    # Renomear colunas do resumo para português
    if 'Status' in df_summary.columns and 'Quantidade' in df_summary.columns and 'Porcentagem' in df_summary.columns:
        pass  # Já está no formato correto
    elif 'Status' in df_summary.columns and 'Count' in df_summary.columns and 'Percentage' in df_summary.columns:
        df_summary = df_summary.rename(columns={
            'Count': 'Quantidade',
            'Percentage': 'Porcentagem'
        })
    
    df_summary.to_excel(writer, sheet_name='Resumo', index=False)
    
    # Converte a lista de elementos HTML para DataFrame
    if isinstance(df_elements, list):
        df_elements = pd.DataFrame(df_elements, columns=['Definition', 'Tag', 'Text', 'Link'])
    
    df_elements.to_excel(writer, sheet_name='Elementos da Página', index=False)
    
    # Salva o arquivo Excel para aplicar formatação
    writer.close()
    
    # Aplicar formatação de cores
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    
    # Formatar aba de comparação
    ws = wb['Comparacao']
    
    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Ajustar largura das colunas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
    
    # Colorir células por status
    color_map = {
        "Exact": "C6EFCE",  # Verde
        "Similar": "FFEB9C",  # Amarelo
        "Partial": "F4B084",  # Laranja
        "Missing": "F8CBAD"   # Vermelho claro
    }
    
    # Encontrar índice da coluna Status
    status_idx = None
    for i, cell in enumerate(ws[1]):
        if cell.value == "Status":
            status_idx = i + 1
            break
    
    if status_idx:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            status = row[status_idx - 1].value
            fill_color = color_map.get(status, "FFFFFF")
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Formatar aba de resumo
    ws = wb['Resumo']
    
    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Ajustar largura das colunas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
    
    # Formatar aba de elementos da página
    ws = wb['Elementos da Página']
    
    # Estilizar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Ajustar largura das colunas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
    
    # Salvar o arquivo com formatação
    wb.save(filename)
    
    return filename
